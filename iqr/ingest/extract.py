"""Deterministic extractors: workbook cells, email structure, doc paragraphs, images.

Pure Python, no model. Every extracted fact carries the leaf's SHA-256 so
citations can anchor to it.
"""
from __future__ import annotations

import email
import email.policy
import io
from pathlib import Path

import openpyxl
from PIL import Image

from iqr.schemas.evidence_graph import (CellFact, DocFact, EmailFact,
                                        EvidenceLeaf, ImageFact)


import os as _os

# tunable for memory-constrained machines; standard workstation defaults are fine
MAX_CELLS_PER_FILE = int(_os.environ.get("IQR_MAX_CELLS_PER_FILE", 4_000_000))
STREAM_THRESHOLD = int(_os.environ.get("IQR_STREAM_THRESHOLD", 60_000_000))


def _extract_workbook_streaming(leaf: EvidenceLeaf) -> tuple[list[CellFact], list[ImageFact]]:
    """Very large workbooks: the dual full load OOMs; stream cached values in
    read_only mode. Formula text and embedded images are unavailable here -
    the truncation sentinel below keeps that honest in the graph."""
    cells: list[CellFact] = []
    wb = openpyxl.load_workbook(str(leaf.stored_path), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None or c.coordinate is None:
                        continue
                    cells.append(CellFact(file_hash=leaf.file_hash, sheet=ws.title,
                                          cell=c.coordinate, value=_plain(c.value)))
                    if len(cells) >= MAX_CELLS_PER_FILE:
                        cells.append(CellFact(file_hash=leaf.file_hash,
                                              sheet="__iqr_meta__", cell="A1",
                                              value=f"TRUNCATED at {MAX_CELLS_PER_FILE} "
                                                    f"cells in sheet {ws.title!r}"))
                        return cells, []
    finally:
        wb.close()
    return cells, []


def extract_workbook(leaf: EvidenceLeaf) -> tuple[list[CellFact], list[ImageFact]]:
    """Extract every non-empty cell (values AND formulas) plus embedded images."""
    if leaf.size > STREAM_THRESHOLD:
        return _extract_workbook_streaming(leaf)
    data = Path(leaf.stored_path).read_bytes()
    cells: list[CellFact] = []
    images: list[ImageFact] = []

    wb_formulas = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    wb_values = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        for row in ws_f.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                raw = c.value
                is_formula = isinstance(raw, str) and raw.startswith("=")
                cached = ws_v[c.coordinate].value if is_formula else raw
                cells.append(CellFact(
                    file_hash=leaf.file_hash, sheet=sheet_name, cell=c.coordinate,
                    value=_plain(cached), is_formula=is_formula,
                    formula=raw if is_formula else None))
        # embedded screenshots (IPE inside workbook tabs)
        for img in getattr(ws_f, "_images", []):
            try:
                blob = img._data()
            except Exception:
                continue
            images.append(_image_fact(blob, f"{leaf.logical_path}!{sheet_name}!{img.ref or 'embedded'}"))
    return cells, images


def _plain(v):
    if v is None or isinstance(v, (int, float, str)):
        return v
    return str(v)


def _image_fact(blob: bytes, logical_path: str) -> ImageFact:
    from iqr.ingest.unpack import sha256, store_blob
    h = sha256(blob)
    stored = store_blob(blob)
    with Image.open(io.BytesIO(blob)) as im:
        w, hgt = im.size
    return ImageFact(file_hash=h, logical_path=logical_path,
                     width=w, height=hgt, stored_path=str(stored))


def extract_email(leaf: EvidenceLeaf, attachment_hashes: list[str]) -> EmailFact:
    data = Path(leaf.stored_path).read_bytes()
    msg = email.message_from_bytes(data, policy=email.policy.default)
    body = msg.get_body(preferencelist=("plain",))
    text = body.get_content() if body else ""
    lines = text.splitlines()
    return EmailFact(
        file_hash=leaf.file_hash,
        message_id=(msg.get("Message-ID") or f"<{leaf.file_hash[:16]}@iqr.local>").strip(),
        sender=msg.get("From", ""),
        to=[a.strip() for a in (msg.get("To", "") or "").split(",") if a.strip()],
        date_raw=msg.get("Date", ""),
        subject=msg.get("Subject", ""),
        lines=lines,
        attachment_hashes=attachment_hashes)


def extract_image(leaf: EvidenceLeaf) -> ImageFact:
    data = Path(leaf.stored_path).read_bytes()
    with Image.open(io.BytesIO(data)) as im:
        w, h = im.size
    return ImageFact(file_hash=leaf.file_hash, logical_path=leaf.logical_path,
                     width=w, height=h, stored_path=str(leaf.stored_path))


W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _stream_docx_paragraphs(path: str) -> list[str]:
    import zipfile
    from lxml import etree
    paras: list[str] = []
    with zipfile.ZipFile(path) as zf, zf.open("word/document.xml") as fh:
        buf: list[str] = []
        for _, el in etree.iterparse(fh, events=("end",)):
            if el.tag == f"{W_NS}t" and el.text:
                buf.append(el.text)
            elif el.tag == f"{W_NS}p":
                text = "".join(buf).strip()
                if text:
                    paras.append(text)
                buf = []
                el.clear()  # free the subtree - this is the point
                while el.getprevious() is not None:
                    del el.getparent()[0]
    return paras


def extract_docx(leaf: EvidenceLeaf) -> tuple[DocFact, list[ImageFact]]:
    """Text paragraphs + every embedded image (screenshot-heavy 404s and
    validation docs carry the IPE as pictures in word/media/)."""
    import zipfile
    if leaf.size > 30_000_000:
        # screenshot-laden 404 docs: python-docx builds the full XML tree and
        # OOMs; stream w:p/w:t nodes instead
        paras = _stream_docx_paragraphs(str(leaf.stored_path))
    else:
        import docx as docx_mod
        d = docx_mod.Document(str(leaf.stored_path))
        paras = [p.text for p in d.paragraphs if p.text.strip()]
    images: list[ImageFact] = []
    with zipfile.ZipFile(str(leaf.stored_path)) as zf:
        for name in sorted(zf.namelist()):
            if name.startswith("word/media/") and not name.endswith("/"):
                blob = zf.read(name)
                if len(blob) < 2048:
                    continue  # icons/separators, not evidence screenshots
                try:
                    images.append(_image_fact(blob, f"{leaf.logical_path}!{name}"))
                except Exception:
                    continue  # emf/wmf and other non-raster media
    return DocFact(file_hash=leaf.file_hash, paragraphs=paras), images


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def extract_xlsb(leaf: EvidenceLeaf) -> list[CellFact]:
    """Binary workbooks via pyxlsb - cached values only (no formula text)."""
    from pyxlsb import open_workbook
    cells: list[CellFact] = []
    with open_workbook(str(leaf.stored_path)) as wb:
        for sheet_name in wb.sheets:
            with wb.get_sheet(sheet_name) as sheet:
                for row in sheet.rows():
                    for item in row:
                        if item.v is None:
                            continue
                        coord = f"{_col_letter(item.c + 1)}{item.r + 1}"
                        cells.append(CellFact(file_hash=leaf.file_hash,
                                              sheet=sheet_name, cell=coord,
                                              value=_plain(item.v)))
    return cells


def extract_msg_email(leaf: EvidenceLeaf, attachment_hashes: list[str]) -> EmailFact:
    """Outlook .msg headers/body via extract-msg."""
    import extract_msg
    msg = extract_msg.openMsg(str(leaf.stored_path))
    try:
        body = msg.body or ""
        lines = body.splitlines()
        mid = (msg.messageId or f"<{leaf.file_hash[:16]}@iqr.local>").strip()
        date_raw = str(msg.date) if msg.date else ""
        return EmailFact(file_hash=leaf.file_hash, message_id=mid,
                         sender=msg.sender or "", to=[t.strip() for t in (msg.to or "").split(";") if t.strip()],
                         date_raw=date_raw, subject=msg.subject or "",
                         lines=lines, attachment_hashes=attachment_hashes)
    finally:
        msg.close()
